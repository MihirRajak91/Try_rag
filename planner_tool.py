import os
from typing import Dict, Any
from openai import OpenAI

from datetime import datetime
from openpyxl import Workbook, load_workbook

from rag.assembler import assemble_prompt
from rag.manifest import write_manifest

LOG_XLSX_PATH = os.getenv("PLANNER_RUNS_XLSX", "manifests/planner_runs.xlsx")

def append_run_to_excel(*, xlsx_path: str, row: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)

    headers = [
        "timestamp",
        "query",
        "winner",
        "topics",
        "secondary",
        "prompt_chars",
        "approx_tokens",
        "llm_output",
    ]

    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "runs"
        ws.append(headers)

    ws.append([
        row.get("timestamp", ""),
        row.get("query", ""),
        row.get("winner", ""),
        row.get("topics", ""),
        row.get("secondary", ""),
        row.get("prompt_chars", ""),
        row.get("approx_tokens", ""),
        row.get("llm_output", ""),
    ])

    wb.save(xlsx_path)

def run_planner_full(query: str, debug: bool = True) -> Dict[str, Any]:
    prompt, audit, manifest = assemble_prompt(
        query,
        debug=debug,
        return_audit=True,
        return_manifest=True,
    )

    manifest_path = write_manifest("manifests", manifest)

    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    resp = client.chat.completions.create(
        model=os.getenv("OPENAI_MODEL", "gpt-4.1-mini"),
        messages=[
            {"role": "system", "content": "You are an expert workflow automation architect."},
            {"role": "user", "content": prompt},
        ],
        temperature=0.2,
    )
    output = resp.choices[0].message.content.strip()

    ts = datetime.now().isoformat(timespec="seconds")

    routing_info = manifest.get("routing", {})
    audit_info = manifest.get("audit", {})

    append_run_to_excel(
        xlsx_path=LOG_XLSX_PATH,
        row={
            "timestamp": ts,
            "query": query,
            "winner": routing_info.get("winner", ""),
            "topics": ", ".join(routing_info.get("topics", []) or []),
            "secondary": ", ".join(routing_info.get("secondary", []) or []),
            "prompt_chars": audit_info.get("prompt_chars", audit.prompt_chars),
            "approx_tokens": audit_info.get("approx_tokens", audit.approx_tokens),
            "llm_output": output,
        },
    )


    return {
        "routing": manifest.get("routing", {}),
        "audit": {
            "prompt_chars": audit.prompt_chars,
            "approx_tokens": audit.approx_tokens,
            "chunks_count": audit.chunks_count,
            "by_role": audit.by_role,
            "by_doc_type": audit.by_doc_type,
            "by_topic": audit.by_topic,
        },
        "manifest_path": manifest_path,
        "manifest": manifest,
        "prompt": prompt,
        "llm_output": output,
    }
